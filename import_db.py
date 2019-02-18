#!/usr/bin/python3
# -*- coding: utf-8 -*-

import os,sys,getopt
import pickle
import pymysql
from openpyxl import load_workbook
from operator import itemgetter, attrgetter

COL_STU_ID = 1
COL_STU_NAME = 2
COL_PINYIN_NAME = 3
COL_ENG_NAME = 4
COL_COURSE_NAME = 5
COL_TEACHER_NAME = 6
COL_CLASS_NAME = 7
COL_FTEACHER_NAME = 8

def is_alphabet_s(ustring):
    for c in ustring:
        if not is_alphabet_c(c):
            return False
    return True;
    
def is_alphabet_c(uchar):
    """ if it's alphabetic char """
    if (uchar >= u'\u0041' and uchar<=u'\u005a') or (uchar >= u'\u0061' and uchar<=u'\u007a'):
        return True
    else:
        return False

def add_teachers(cursor,tindex,set_t,special=None):
    pindex = 0
    for t_name,t_tel in set_t:
        if special != None and t_name == special:
            pindex = tindex
            tindex = 1
        elif tindex == 1:
            tindex = tindex+1
        cursor.execute('insert into teachers(id,name,phone) values(%s,%s,%s)',(tindex,t_name.encode('utf-8'),t_tel))
        if pindex !=0 :
            tindex = pindex
            pindex = 0
        else:
            tindex = tindex +1
        
def import_db_teachers(db,ws):
    # insert into teachers,classes
    print('Importing teachers standby...')
    COL_NAME = 1
    COL_TEL = 2
    COL_TYPE = 3
    COL_CLASS1 = 4
    COL_CLASS2 = 5
    with db.cursor() as cursor:
        cursor.execute('set foreign_key_checks = 0')
        cursor.execute('truncate teachers')
        cursor.execute('truncate classes')
        cursor.execute('set foreign_key_checks = 1')

        native_idx=1
        foreign_idx=1000
        form_idx=200
        other_idx=300
        class_ft_map = {}
        t_count = 0
        max_row = ws.max_row+1 
        for i in range(2,max_row):
            t_name = ws.cell(row=i,column=COL_NAME).value
            t_tel = ws.cell(row=i,column=COL_TEL).value
            t_type_val = ws.cell(row=i,column=COL_TYPE).value
            t_class1 = ws.cell(row=i,column=COL_CLASS1).value
            t_class2 = ws.cell(row=i,column=COL_CLASS2).value
            if(t_name != None):
                t_name = t_name.strip()
                t_type = int(t_type_val) if t_type_val else 0
                if t_type == 0:
                    index = native_idx
                    native_idx += 1
                elif t_type == 1:
                    index = foreign_idx
                    foreign_idx += 1
                elif t_type == 2:
                    index = form_idx
                    form_idx += 1
                    if t_class1 != None:
                        class_ft_map.update({t_class1:index})
                    if t_class2 != None:
                        class_ft_map.update({t_class2:index})
                else:
                    index = other_idx
                    other_idx +=1 

                t_count += 1
                # insert into teachers
                cursor.execute('insert into teachers(id,name,phone,type) values(%s,%s,%s,%s)',
                               (index,t_name.encode('utf-8'),t_tel,t_type))

        # insert into classes
        for (cl_name,t_id) in class_ft_map.items():
            cursor.execute('insert into classes(name,ft_id) values(%s,%s)',(cl_name,t_id))
            
        db.commit()
                
        print('Importing teachers done! count=%d'%t_count)

def get_name_id_map(cursor,sql):
    cursor.execute(sql)
    data_tbl = cursor.fetchall()
    data_map = {}
    for id,name in data_tbl:
        data_map[name.strip()] = id

    return data_map

def format_course_name(cname):
    cname_list = cname.split()
    return ' '.join(cname_list),cname_list

def import_db_timetable(db,ws):
    # insert into courses, timetable
    print('Importing timetable standby...')
    COL_DAY = 1
    COL_SLOT = 2
    COL_COURSE = 3
    COL_TEACHER = 4
    COL_CLASSROOM = 5
    
    day_map = {'MON':1,'TUE':2,'WED':3,'THU':4,'FRI':5,'SAT':6,'SUN':7}
    with db.cursor() as cursor:
        cursor.execute('set foreign_key_checks = 0')
        cursor.execute('truncate courses')
        cursor.execute('truncate timetable')
        cursor.execute('set foreign_key_checks = 1')

        t_name_id_map = get_name_id_map(cursor, 'select id,name from teachers')
        cr_name_id_map = get_name_id_map(cursor, 'select id,name from classrooms')
        
        course_id_map = {}
        course_idx = 0
        course_id = 0
        max_row = ws.max_row+1 
        for i in range(2,max_row):
            col_day = ws.cell(row=i,column=COL_DAY).value
            col_slot = ws.cell(row=i,column=COL_SLOT).value
            c_name = ws.cell(row=i,column=COL_COURSE).value
            t_name = ws.cell(row=i,column=COL_TEACHER).value
            cr_name = ws.cell(row=i,column=COL_CLASSROOM).value
            
            if (c_name == None or t_name==None or  cr_name==None or col_day==None or col_slot==None):
                print('Timetable: incomplete info. row=%d.'%i)
                continue

            tt_slot = int(col_slot)
            tt_day = day_map.get(col_day.upper(),0)
            if tt_day == 0:
                print(('Timetable: unknown day,row=%d,week=%s')%(i,col_day))
                continue
            
            t_name = t_name.strip()
            t_id = t_name_id_map.get(t_name,0)
            if t_id == 0:
                print(('Timetable: unknown teacher,row=%d,name=%s')%(i,t_name))
                continue
                
            cr_name = cr_name.strip()
            cr_id = cr_name_id_map.get(cr_name,0)
            if cr_id == 0:
                print(('Timetable: unknown classroom,row=%d,room=%s')%(i,cr_name))
                continue            
            
            c_name = c_name.strip()                
            course_id = course_id_map.get(c_name,0)
            if (course_id == 0):
                # new course, add this course into db and set
                course_idx += 1
                course_id = course_idx
                course_id_map.update({c_name:course_idx})
                cursor.execute('insert into courses(id,sname,fname) values(%s,%s,%s)',
                               (course_idx,c_name.encode('utf-8'),c_name.encode('utf-8')))
                               
            # insert into timetable
            cursor.execute('insert into timetable (day,slot,t_id,c_id,cr_id) values (%s,%s,%s,%s,%s)',
                           (tt_day, tt_slot, t_id, course_id, cr_id))

        db.commit()
  
    print('Importing timetable done,courses=%d!'%course_idx)

    
def import_db_enrollment(db,ws):
    # insert into students
    # insert into enrollment

    COL_CLASS = 1
    COL_STU_ID = 2
    COL_STU_NAME = 3
    COL_STU_PINYIN = 4
    COL_STU_ENG = 5
    COL_STU_GENDER = 6
    COL_COURSE = 8
    
    print('Importing enrollment standby...')
    with db.cursor() as cursor:
        # clear students and enrollment
        cursor.execute('set foreign_key_checks = 0')
        cursor.execute('truncate students')
        cursor.execute('truncate enrollment')
        cursor.execute('set foreign_key_checks = 1')
        
        course_name_map = get_name_id_map(cursor,'select id,sname from courses')
        class_name_map = get_name_id_map(cursor,'select id,name from classes')

        # course and teacher id mapping, one course may have multiple teachers
        cursor.execute('select c_id,t_id from timetable group by t_id,c_id order by c_id')
        tt_ct_tbl = cursor.fetchall()
        course_teacher_id_map = {}
        for c_id,t_id in tt_ct_tbl:
            t_ids = course_teacher_id_map.get(c_id,None)
            if t_ids == None:
                course_teacher_id_map.update({c_id:[t_id]})
            else:
                t_ids.append(t_id)
                course_teacher_id_map.update({c_id:t_ids})

        stu_id_map = set()
        stu_index = 0
        e_count = 0
        
        max_row = ws.max_row+1
        for i in range(2,max_row):
            s_class_name = ws.cell(row=i,column=COL_CLASS).value
            s_id = ws.cell(row=i,column=COL_STU_ID).value
            s_name = ws.cell(row=i,column=COL_STU_NAME).value
            s_pinyin = ws.cell(row=i,column=COL_STU_PINYIN).value
            s_eng = ws.cell(row=i,column=COL_STU_ENG).value
            col_gender = ws.cell(row=i,column=COL_STU_GENDER).value

            s_id = s_id.strip()
            if s_id in stu_id_map:
                print('Enrollment: duplicate student id=%s,line=%d'%(s_id,i))
                continue
            else:
                stu_id_map.add(s_id)

            cl_id = class_name_map.get(s_class_name.strip(),0)
            if(cl_id == 0):
                print('Enrollment: unkonw class=%s,line=%d'%(s_class_name.strip(),i))
                continue

            # new student info
            s_gender = 1 if (col_gender.strip().upper()=='M') else 0
            stu_index += 1
            cursor.execute('insert into students(id,stu_id,name,py_name,eng_name,gender,cl_id) values(%s,%s,%s,%s,%s,%s,%s)',
                           (stu_index,s_id,s_name.encode('utf-8'),s_pinyin,s_eng,s_gender,cl_id))
            

            s_courses_id=[]
            for j in range(0,5):
                course_name = ws.cell(row=i,column=COL_COURSE+j).value
                course_id = course_name_map.get(course_name.strip(),0)
                if(course_id != 0):
                    s_courses_id.append(course_id)
                else:
                    print('Enrollment: unknown course=%s,line=%d'%(course_name,i))

            if len(s_courses_id) != 5:
                continue

            for c_id in s_courses_id:
                t_ids = course_teacher_id_map.get(c_id)
                if t_ids == None:
                    print('Enrollment: no teacher for this c_id=%d,line=%d'%(c_id,i))
                    continue
                for t_id in t_ids:
                    cursor.execute('insert into enrollment(s_id,t_id,c_id) values(%s,%s,%s)',(stu_index,t_id,c_id))
                    e_count += 1
        
        db.commit()

        print('Importing %d students done!'%stu_index)
        print('Importing %d enrollment done!'%e_count)

def setup_db(xlsx_file):
    db = pymysql.connect(user='vw', password='letmein', database='hs_timetable', charset='utf8')
    wb = load_workbook(xlsx_file)
    ### ws = wb.active  
    ws = wb['Teachers']
    import_db_teachers(db,ws)
    ws = wb['Timetable']
    import_db_timetable(db,ws)
    ws = wb['Enrollment']
    import_db_enrollment(db,ws)
    db.close()

    '''
db_hs_config = {
    'user': 'vw',
    'password': 'letmein',
    'database': 'hs_timetable',
    'charset': 'utf8'
}    
def db_conn_pool():
    # * means put para into a tuple
    # ** means put para into a dict
    pool = PyMysqlPool.ConnectionPool(**db_hs_config)
    pool.connect()
    return pool
    
def test_db_pool():
    with db_conn_pool().cursor as cursor:
        cursor.execute('select * from student')
        print(cursor.fetchone())

        counter = 0
        for stu in cursor:
            print(stu)
            counter = counter+1
            if counter > 10:
                break
                
'''
def main():
    try:
        opts,args = getopt.getopt(sys.argv[1:],'')
    except:
        print('Usage:')
        print('  import_db.py file.xlsx')
        return

    yes_or_no = input("Continue to clear the db and import data, Y/N? ")
    if yes_or_no == "Y":
        setup_db(args[0])
    else:
        print("Abort action!")


if __name__ == '__main__':
    main()
